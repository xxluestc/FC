"""Build auditable one-day/three-day dual-stack datasets and figures.

No vehicle-demand or battery-power semantics are assumed. The script first
tests the raw ``target power`` field against measured fuel-cell and DCDC power.
"""

from __future__ import annotations

from datetime import timedelta
from pathlib import Path
import re

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[3] / "2026\u674e\u4fca\u8c6a" / "\u5b66\u4f4d\u8bba\u6587" / "\u5b9e\u9a8c\u6570\u636e"
OUT = Path("experiments/li_junhao_dual_stack_audit")
ENGINES = {
    "\u4e09\u53f7\u53d1\u52a8\u673a": "all_data",
    "\u56db\u53f7\u53d1\u52a8\u673a": "YF20YF21-20210707-20220110",
    "\u4e94\u53f7\u53d1\u52a8\u673a": "YF24-20210908-20220208",
}
SYSTEM_FIELDS = [
    "\u65f6\u95f4", "\u5de5\u4f5c\u72b6\u6001", "\u7535\u6c60\u6545\u969c\u7b49\u7ea7", "DCDC\u8f93\u51fa\u529f\u7387",
    "\u71c3\u6599\u7535\u6c60\u7535\u538b", "\u71c3\u6599\u7535\u6c60\u7535\u6d41", "\u6c22\u71c3\u6599\u7535\u6c60\u8f93\u51fa\u529f\u7387",
    "A\u5806\u7535\u538b\u5e73\u5747\u503c", "B\u5806\u7535\u538b\u5e73\u5747\u503c", "\u7a7a\u6c14\u6d41\u91cf",
    "\u6c34\u8fdb\u5806\u6e29\u5ea6", "\u6c34\u51faA\u5806\u6e29\u5ea6", "B\u5806\u6c34\u51fa\u5806\u6e29\u5ea6", "Purge\u6807\u5fd7\u4f4d",
]
VEHICLE_FIELDS = [
    "\u65f6\u95f4", "\u76ee\u6807\u529f\u7387", "\u8f66\u901f", "\u7535\u6c60SOC", "\u7d2f\u8ba1\u884c\u9a76\u516c\u91cc\u6570",
    "\u5f53\u524d\u884c\u9a76\u516c\u91cc\u6570", "\u6c22\u6c14\u5269\u4f59\u8d28\u91cf", "\u6c22\u6c14\u6d88\u8017\u7387", "SOP\u53ef\u52a0\u8f7d\u529f\u7387",
]
RENAME = {
    "\u65f6\u95f4": "timestamp", "\u5de5\u4f5c\u72b6\u6001": "state", "\u7535\u6c60\u6545\u969c\u7b49\u7ea7": "fault_level",
    "DCDC\u8f93\u51fa\u529f\u7387": "dcdc_power_kW", "\u71c3\u6599\u7535\u6c60\u7535\u538b": "fc_voltage_V",
    "\u71c3\u6599\u7535\u6c60\u7535\u6d41": "fc_current_A", "\u6c22\u71c3\u6599\u7535\u6c60\u8f93\u51fa\u529f\u7387": "fc_power_kW",
    "A\u5806\u7535\u538b\u5e73\u5747\u503c": "stack_A_section_avg_V", "B\u5806\u7535\u538b\u5e73\u5747\u503c": "stack_B_section_avg_V",
    "\u7a7a\u6c14\u6d41\u91cf": "airflow", "\u6c34\u8fdb\u5806\u6e29\u5ea6": "water_in_C", "\u6c34\u51faA\u5806\u6e29\u5ea6": "stack_A_water_out_C",
    "B\u5806\u6c34\u51fa\u5806\u6e29\u5ea6": "stack_B_water_out_C", "Purge\u6807\u5fd7\u4f4d": "purge_flag",
    "\u76ee\u6807\u529f\u7387": "target_power_unknown_kW", "\u8f66\u901f": "vehicle_speed", "\u7535\u6c60SOC": "battery_soc",
    "\u7d2f\u8ba1\u884c\u9a76\u516c\u91cc\u6570": "odometer_km", "\u5f53\u524d\u884c\u9a76\u516c\u91cc\u6570": "trip_km",
    "\u6c22\u6c14\u5269\u4f59\u8d28\u91cf": "hydrogen_remaining", "\u6c22\u6c14\u6d88\u8017\u7387": "hydrogen_consumption_rate",
    "SOP\u53ef\u52a0\u8f7d\u529f\u7387": "sop_available_power_kW",
}


def style() -> None:
    plt.rcParams.update({"font.sans-serif": ["Microsoft YaHei", "SimHei"], "axes.unicode_minus": False,
                         "figure.dpi": 130, "savefig.dpi": 320, "axes.grid": True, "grid.alpha": .22,
                         "axes.spines.top": False, "axes.spines.right": False})


def date_of(path: Path) -> pd.Timestamp:
    return pd.Timestamp(re.search(r"_(\d{8})", path.name).group(1))


def select_triple(engine: str, folder: str) -> list[Path]:
    files = [p for p in sorted((ROOT / folder).glob(f"{engine}_*.xlsx")) if p.stat().st_size > 1_000_000]
    dated = {date_of(p): p for p in files}
    candidates = []
    for day, path in dated.items():
        trio = [dated.get(day + timedelta(days=offset)) for offset in range(3)]
        if all(trio):
            candidates.append((sum(p.stat().st_size for p in trio), trio))
    if not candidates:
        raise RuntimeError(f"No contiguous three-day block: {engine}")
    return max(candidates, key=lambda item: item[0])[1]


def read_sheet(path: Path, sheet: str, requested: list[str]) -> pd.DataFrame:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    worksheet = workbook[sheet]
    iterator = worksheet.iter_rows(values_only=True)
    headers = list(next(iterator))
    index = {str(value).strip(): position for position, value in enumerate(headers) if value is not None}
    available = [name for name in requested if name in index]
    rows = [[row[index[name]] for name in available] for row in iterator]
    workbook.close()
    return pd.DataFrame(rows, columns=available).rename(columns=RENAME)


def load_day(path: Path) -> pd.DataFrame:
    system = read_sheet(path, "\u7cfb\u7edf\u603b\u89c8", SYSTEM_FIELDS)
    vehicle = read_sheet(path, "\u6574\u8f66\u72b6\u6001", VEHICLE_FIELDS)
    for frame in (system, vehicle):
        frame["timestamp"] = pd.to_datetime(frame["timestamp"], errors="coerce")
        frame.dropna(subset=["timestamp"], inplace=True)
        frame.sort_values("timestamp", inplace=True)
        frame.drop_duplicates("timestamp", keep="last", inplace=True)
    merged = pd.merge_asof(system, vehicle, on="timestamp", direction="nearest", tolerance=pd.Timedelta(seconds=1))
    numeric_columns = [column for column in merged.columns if column not in {"timestamp", "state"}]
    merged[numeric_columns] = merged[numeric_columns].apply(pd.to_numeric, errors="coerce")
    merged["source_file"] = path.name
    merged["source_date"] = date_of(path)
    # The fifth engine has 70 exported cell-voltage columns but its overview
    # average represents about 56 effective voltage sections per stack. Infer
    # the effective count from the independently measured total stack voltage.
    section_ratio = merged["fc_voltage_V"] / (
        merged["stack_A_section_avg_V"] + merged["stack_B_section_avg_V"]
    )
    effective_sections = section_ratio.replace([np.inf, -np.inf], np.nan).median()
    merged["effective_sections_per_stack"] = effective_sections
    voltage_sum = merged["stack_A_section_avg_V"] + merged["stack_B_section_avg_V"]
    # Allocate independently measured total voltage by the A/B average-voltage
    # ratio. This avoids imposing 70 active sections on the fifth engine.
    merged["stack_A_voltage_V"] = merged["fc_voltage_V"] * merged["stack_A_section_avg_V"] / voltage_sum
    merged["stack_B_voltage_V"] = merged["fc_voltage_V"] * merged["stack_B_section_avg_V"] / voltage_sum
    merged["stack_A_power_kW_inferred"] = merged["stack_A_voltage_V"] * merged["fc_current_A"] / 1000
    merged["stack_B_power_kW_inferred"] = merged["stack_B_voltage_V"] * merged["fc_current_A"] / 1000
    merged["stack_power_sum_kW_inferred"] = merged["stack_A_power_kW_inferred"] + merged["stack_B_power_kW_inferred"]
    merged["target_minus_fc_kW_unverified"] = merged["target_power_unknown_kW"] - merged["fc_power_kW"]
    return merged


def compressed_time(frame: pd.DataFrame) -> np.ndarray:
    delta = frame["timestamp"].diff().dt.total_seconds().fillna(0).clip(lower=0, upper=5)
    return delta.cumsum().to_numpy() / 3600


def thin(frame: pd.DataFrame, maximum: int = 6000) -> pd.DataFrame:
    stride = max(1, len(frame) // maximum)
    return frame.iloc[::stride].copy()


def valid_operation(frame: pd.DataFrame) -> pd.DataFrame:
    mask = frame["fc_current_A"].between(5, 350) & frame["fc_voltage_V"].between(80, 300)
    if "state" in frame:
        mask &= frame["state"].astype(str).str.contains("\u8fd0\u884c", na=False)
    if "fault_level" in frame:
        mask &= frame["fault_level"].fillna(0).eq(0)
    return frame.loc[mask].sort_values("timestamp").reset_index(drop=True)


def plot_stack_power(frame: pd.DataFrame, engine: str, label: str, path: Path) -> None:
    data = thin(valid_operation(frame))
    x = compressed_time(data)
    fig, axes = plt.subplots(2, 1, figsize=(10.5, 7.2), sharex=True)
    axes[0].plot(x, data["stack_A_power_kW_inferred"], lw=.9, label="A\u5806\u63a8\u7b97\u529f\u7387", color="#0072B2")
    axes[0].plot(x, data["stack_B_power_kW_inferred"], lw=.9, label="B\u5806\u63a8\u7b97\u529f\u7387", color="#D55E00")
    axes[0].plot(x, data["fc_power_kW"], lw=.8, alpha=.65, label="\u5b9e\u6d4bFC\u603b\u529f\u7387", color="#333333")
    total = data["stack_power_sum_kW_inferred"].replace(0, np.nan)
    axes[1].plot(x, 100 * data["stack_A_power_kW_inferred"] / total, lw=.9, label="A\u5806\u529f\u7387\u4efd\u989d", color="#0072B2")
    axes[1].plot(x, 100 * data["stack_B_power_kW_inferred"] / total, lw=.9, label="B\u5806\u529f\u7387\u4efd\u989d", color="#D55E00")
    axes[0].set_ylabel("\u529f\u7387 / kW"); axes[1].set_ylabel("\u529f\u7387\u4efd\u989d / %"); axes[1].set_xlabel("\u538b\u7f29\u540e\u8fd0\u884c\u65f6\u95f4 / h")
    axes[0].legend(ncol=3, fontsize=8, frameon=False); axes[1].legend(ncol=2, fontsize=8, frameon=False)
    fig.suptitle(f"{engine} {label}\uff1a\u4e32\u8054A/B\u5806\u7535\u529f\u7387\u8d21\u732e")
    fig.tight_layout(); fig.savefig(path, bbox_inches="tight"); plt.close(fig)


def plot_power_semantics(frame: pd.DataFrame, engine: str, label: str, path: Path) -> None:
    data = thin(frame.dropna(subset=["target_power_unknown_kW", "fc_power_kW"]).sort_values("timestamp").reset_index(drop=True))
    x = compressed_time(data)
    fig, axes = plt.subplots(2, 1, figsize=(10.5, 7.2), sharex=True)
    axes[0].plot(x, data["target_power_unknown_kW"], lw=.9, label="\u76ee\u6807\u529f\u7387\uff08\u8bed\u4e49\u5f85\u5224\u5b9a\uff09", color="#CC79A7")
    axes[0].plot(x, data["fc_power_kW"], lw=.85, label="\u71c3\u6599\u7535\u6c60\u539f\u59cb\u8f93\u51fa", color="#0072B2")
    axes[0].plot(x, data["dcdc_power_kW"], lw=.8, label="DCDC\u8f93\u51fa", color="#E69F00")
    axes[1].plot(x, data["target_minus_fc_kW_unverified"], lw=.8, color="#555555", label="\u76ee\u6807-FC\u6b8b\u5dee\uff08\u4e0d\u7b49\u540c\u7535\u6c60\u529f\u7387\uff09")
    axes[0].set_ylabel("\u529f\u7387 / kW"); axes[1].set_ylabel("\u6b8b\u5dee / kW"); axes[1].set_xlabel("\u538b\u7f29\u540e\u65f6\u95f4 / h")
    axes[0].legend(ncol=3, fontsize=8, frameon=False); axes[1].legend(frameon=False, fontsize=8)
    fig.suptitle(f"{engine} {label}\uff1a\u529f\u7387\u5b57\u6bb5\u5173\u7cfb\u5ba1\u8ba1")
    fig.tight_layout(); fig.savefig(path, bbox_inches="tight"); plt.close(fig)


def metrics(frame: pd.DataFrame, engine: str, selected: list[Path]) -> dict:
    data = frame.dropna(subset=["target_power_unknown_kW", "fc_power_kW", "dcdc_power_kW"])
    positive = data[data["fc_power_kW"] > 1]
    stack = valid_operation(frame).dropna(subset=["stack_power_sum_kW_inferred", "fc_power_kW"])
    def compare(a: str, b: str) -> tuple[float, float, float]:
        pair = data[[a, b]].dropna()
        return pair[a].corr(pair[b]), (pair[a] - pair[b]).abs().mean(), np.sqrt(np.mean((pair[a] - pair[b]) ** 2))
    t_fc = compare("target_power_unknown_kW", "fc_power_kW")
    t_dc = compare("target_power_unknown_kW", "dcdc_power_kW")
    return {
        "engine": engine, "three_day_files": " | ".join(path.name for path in selected),
        "rows": len(frame), "matched_target_rows": len(data),
        "target_fc_corr": t_fc[0], "target_fc_mae_kW": t_fc[1], "target_fc_rmse_kW": t_fc[2],
        "target_dcdc_corr": t_dc[0], "target_dcdc_mae_kW": t_dc[1], "target_dcdc_rmse_kW": t_dc[2],
        "median_target_to_fc_ratio_when_fc_gt1": np.nanmedian(positive["target_power_unknown_kW"] / positive["fc_power_kW"]),
        "stack_sum_fc_corr": stack["stack_power_sum_kW_inferred"].corr(stack["fc_power_kW"]),
        "stack_sum_fc_mae_kW": (stack["stack_power_sum_kW_inferred"] - stack["fc_power_kW"]).abs().mean(),
        "hydrogen_rate_nonmissing_pct": 100 * frame["hydrogen_consumption_rate"].notna().mean(),
        "speed_nonmissing_pct": 100 * frame["vehicle_speed"].notna().mean(),
        "soc_nonmissing_pct": 100 * frame["battery_soc"].notna().mean(),
    }


def plot_ab_decay(engine: str, source: pd.DataFrame, path: Path) -> None:
    data = source[(source["engine"] == engine) & (source["point_count"] >= 10)].copy()
    fig, axes = plt.subplots(2, 1, figsize=(10, 8.2), sharex=True)
    colors = plt.cm.viridis(np.linspace(.08, .92, data["current_point_A"].nunique()))
    section_count = 56 if engine == "\u4e94\u53f7\u53d1\u52a8\u673a" else 70
    for color, (current, group) in zip(colors, data.groupby("current_point_A")):
        group = group.sort_values("date")
        for axis, column, stack in [(axes[0], "stack_A_avg_median", "A"), (axes[1], "stack_B_avg_median", "B")]:
            voltage = section_count * group[column]
            axis.plot(pd.to_datetime(group["date"]), voltage, "o-", color=color, lw=1.5, ms=3.5, label=f"{current:g} A")
            axis.set_ylabel(f"{stack}\u5806\u63a8\u7b97\u5806\u7535\u538b / V")
    axes[0].set_title("A\u5806"); axes[1].set_title("B\u5806"); axes[1].set_xlabel("\u65e5\u671f")
    axes[0].legend(ncol=6, fontsize=8, frameon=False); axes[1].legend(ncol=6, fontsize=8, frameon=False)
    fig.suptitle(f"{engine}\uff1aA/B\u5806\u5404\u5178\u578b\u7535\u6d41\u70b9\u7535\u538b\u8d8b\u52bf")
    fig.tight_layout(); fig.savefig(path, bbox_inches="tight"); plt.close(fig)


def main() -> None:
    style(); OUT.mkdir(parents=True, exist_ok=True); plots = OUT / "plots"; plots.mkdir(exist_ok=True)
    metric_rows = []
    selections = []
    for engine, folder in ENGINES.items():
        selected = select_triple(engine, folder)
        selections.extend({"engine": engine, "scope": "three_day", "file": path.name, "size_MB": path.stat().st_size / 1e6} for path in selected)
        print(engine, [path.name for path in selected], flush=True)
        frames = [load_day(path) for path in selected]
        three_day = pd.concat(frames, ignore_index=True).sort_values("timestamp").reset_index(drop=True)
        one_path = max(selected, key=lambda path: path.stat().st_size)
        one_day = three_day[three_day["source_file"] == one_path.name].copy().reset_index(drop=True)
        selections.append({"engine": engine, "scope": "one_day", "file": one_path.name, "size_MB": one_path.stat().st_size / 1e6})
        three_day.to_csv(OUT / f"{engine}_three_day_processed.csv", index=False, encoding="utf-8-sig")
        metric_rows.append(metrics(three_day, engine, selected))
        plot_stack_power(one_day, engine, "1\u5929", plots / f"{engine}_1day_dual_stack_power.png")
        plot_stack_power(three_day, engine, "\u8fde\u7eed3\u5929", plots / f"{engine}_3day_dual_stack_power.png")
        plot_power_semantics(one_day, engine, "1\u5929", plots / f"{engine}_1day_power_field_audit.png")
        plot_power_semantics(three_day, engine, "\u8fde\u7eed3\u5929", plots / f"{engine}_3day_power_field_audit.png")
    pd.DataFrame(selections).to_csv(OUT / "selected_days.csv", index=False, encoding="utf-8-sig")
    pd.DataFrame(metric_rows).to_csv(OUT / "power_relationship_metrics.csv", index=False, encoding="utf-8-sig")
    decay_source = Path("experiments/li_junhao_voltage_audit/representative_day_voltage_summary.csv")
    if decay_source.exists():
        decay = pd.read_csv(decay_source)
        for engine in ENGINES:
            plot_ab_decay(engine, decay, plots / f"{engine}_AB_voltage_decay.png")


if __name__ == "__main__":
    main()
