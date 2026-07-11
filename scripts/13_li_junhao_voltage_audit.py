"""Audit Li Junhao's three-engine voltage data and plot typical-current trends.

The source workbooks are very large. This audit therefore uses dates sampled
uniformly over each engine's full archive. It only reads columns required for
the voltage audit and never modifies the source data.
"""

from __future__ import annotations

import argparse
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import math
import re

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook


ENGINE_DIRS = {
    "\u4e09\u53f7\u53d1\u52a8\u673a": "all_data",
    "\u56db\u53f7\u53d1\u52a8\u673a": "YF20YF21-20210707-20220110",
    "\u4e94\u53f7\u53d1\u52a8\u673a": "YF24-20210908-20220208",
}
DEFAULT_DATA_ROOT = (
    Path(__file__).resolve().parents[3]
    / "2026\u674e\u4fca\u8c6a"
    / "\u5b66\u4f4d\u8bba\u6587"
    / "\u5b9e\u9a8c\u6570\u636e"
)
TYPICAL_CURRENTS = {
    "\u4e09\u53f7\u53d1\u52a8\u673a": [40, 65, 90, 120, 150],
    "\u56db\u53f7\u53d1\u52a8\u673a": [90, 120, 150, 170],
    "\u4e94\u53f7\u53d1\u52a8\u673a": [55, 90, 120, 160, 195, 220],
}
SHEET = "\u7cfb\u7edf\u603b\u89c8"
COLUMNS = {
    "time": "\u65f6\u95f4",
    "state": "\u5de5\u4f5c\u72b6\u6001",
    "fault": "\u7535\u6c60\u6545\u969c\u7b49\u7ea7",
    "voltage": "\u71c3\u6599\u7535\u6c60\u7535\u538b",
    "current": "\u71c3\u6599\u7535\u6c60\u7535\u6d41",
    "stack_a": "A\u5806\u7535\u538b\u5e73\u5747\u503c",
    "stack_b": "B\u5806\u7535\u538b\u5e73\u5747\u503c",
}


@dataclass
class FileRecord:
    engine: str
    path: Path
    date: pd.Timestamp


def configure_plot_style() -> None:
    plt.rcParams.update(
        {
            "font.sans-serif": ["Microsoft YaHei", "SimHei", "Arial Unicode MS"],
            "axes.unicode_minus": False,
            "figure.dpi": 130,
            "savefig.dpi": 320,
            "axes.grid": True,
            "grid.alpha": 0.22,
            "axes.spines.top": False,
            "axes.spines.right": False,
        }
    )


def parse_date(name: str) -> pd.Timestamp | None:
    match = re.search(r"_(\d{8})", name)
    return pd.Timestamp(datetime.strptime(match.group(1), "%Y%m%d")) if match else None


def inventory(data_root: Path) -> tuple[list[FileRecord], pd.DataFrame]:
    records: list[FileRecord] = []
    rows: list[dict] = []
    for engine, relative in ENGINE_DIRS.items():
        directory = data_root / relative
        candidates = sorted(directory.glob("*.xlsx"))
        valid = []
        invalid = []
        for path in candidates:
            date = parse_date(path.name)
            if path.name.startswith("~$") or path.stat().st_size == 0 or date is None:
                invalid.append(path)
                continue
            if engine not in path.name:
                invalid.append(path)
                continue
            valid.append(FileRecord(engine, path, date))
        records.extend(valid)
        rows.append(
            {
                "engine": engine,
                "directory": relative,
                "xlsx_total": len(candidates),
                "valid_daily_files": len(valid),
                "excluded_files": len(invalid),
                "start_date": min(item.date for item in valid).date(),
                "end_date": max(item.date for item in valid).date(),
                "archive_size_GB": sum(item.path.stat().st_size for item in valid) / 1e9,
            }
        )
    return records, pd.DataFrame(rows)


def uniform_sample(records: list[FileRecord], count: int) -> list[FileRecord]:
    if len(records) <= count:
        return records
    indices = np.linspace(0, len(records) - 1, count).round().astype(int)
    return [records[index] for index in sorted(set(indices))]


def numeric(value: object) -> float | None:
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if math.isfinite(result) else None


def read_daily_summary(record: FileRecord, tolerance: float) -> list[dict]:
    workbook = load_workbook(record.path, read_only=True, data_only=True)
    if SHEET not in workbook.sheetnames:
        workbook.close()
        return [{"engine": record.engine, "date": record.date, "status": "missing_sheet"}]
    worksheet = workbook[SHEET]
    rows = worksheet.iter_rows(values_only=True)
    headers = list(next(rows))
    indices = {str(value).strip(): index for index, value in enumerate(headers) if value is not None}
    required = [COLUMNS["voltage"], COLUMNS["current"]]
    if any(column not in indices for column in required):
        workbook.close()
        return [{"engine": record.engine, "date": record.date, "status": "missing_columns"}]

    voltage_values: list[float] = []
    current_values: list[float] = []
    by_current: dict[int, dict[str, list[float]]] = {
        center: defaultdict(list) for center in TYPICAL_CURRENTS[record.engine]
    }
    state_index = indices.get(COLUMNS["state"])
    fault_index = indices.get(COLUMNS["fault"])
    voltage_index = indices[COLUMNS["voltage"]]
    current_index = indices[COLUMNS["current"]]
    stack_a_index = indices.get(COLUMNS["stack_a"])
    stack_b_index = indices.get(COLUMNS["stack_b"])

    for row in rows:
        current = numeric(row[current_index])
        voltage = numeric(row[voltage_index])
        if current is None or voltage is None or not (5 < current < 350 and 80 < voltage < 300):
            continue
        if state_index is not None and "\u8fd0\u884c" not in str(row[state_index]).strip():
            continue
        fault = numeric(row[fault_index]) if fault_index is not None else 0
        if fault not in (None, 0.0):
            continue
        voltage_values.append(voltage)
        current_values.append(current)
        for center in by_current:
            if abs(current - center) <= tolerance:
                by_current[center]["voltage"].append(voltage)
                stack_a = numeric(row[stack_a_index]) if stack_a_index is not None else None
                stack_b = numeric(row[stack_b_index]) if stack_b_index is not None else None
                if stack_a is not None:
                    by_current[center]["stack_a"].append(stack_a)
                if stack_b is not None:
                    by_current[center]["stack_b"].append(stack_b)
                break
    workbook.close()

    output: list[dict] = []
    common = {
        "engine": record.engine,
        "date": record.date,
        "file": record.path.name,
        "status": "ok" if voltage_values else "no_valid_rows",
        "operating_count": len(voltage_values),
        "overall_voltage_median_V": np.median(voltage_values) if voltage_values else np.nan,
        "overall_current_median_A": np.median(current_values) if current_values else np.nan,
    }
    for center, values in by_current.items():
        voltages = values["voltage"]
        output.append(
            {
                **common,
                "current_point_A": center,
                "point_count": len(voltages),
                "voltage_median_V": np.median(voltages) if voltages else np.nan,
                "voltage_p10_V": np.quantile(voltages, 0.10) if voltages else np.nan,
                "voltage_p90_V": np.quantile(voltages, 0.90) if voltages else np.nan,
                "stack_A_avg_median": np.median(values["stack_a"]) if values["stack_a"] else np.nan,
                "stack_B_avg_median": np.median(values["stack_b"]) if values["stack_b"] else np.nan,
            }
        )
    return output or [common]


def add_filter(frame: pd.DataFrame, column: str) -> pd.Series:
    return frame[column].rolling(window=3, center=True, min_periods=1).median()


def plot_overall(frame: pd.DataFrame, output: Path) -> None:
    fig, axis = plt.subplots(figsize=(10.5, 5.7))
    colors = ["#2166AC", "#B2182B", "#1B7837"]
    for (engine, group), color in zip(frame.groupby("engine", sort=False), colors):
        daily = group.drop_duplicates(["engine", "date"]).sort_values("date")
        axis.plot(daily["date"], daily["overall_voltage_median_V"], "o-", color=color, alpha=0.38, lw=1, ms=3.8, label=f"{engine} 原始")
        axis.plot(daily["date"], add_filter(daily, "overall_voltage_median_V"), color=color, lw=2.5, label=f"{engine} 滤波")
    axis.set(title="三台燃料电池发动机运行电压趋势（代表日）", xlabel="日期", ylabel="系统总电压日中位数 / V")
    axis.legend(ncol=2, fontsize=8, frameon=False)
    fig.tight_layout()
    fig.savefig(output, bbox_inches="tight")
    plt.close(fig)


def plot_typical(frame: pd.DataFrame, output: Path) -> None:
    engines = list(ENGINE_DIRS)
    fig, axes = plt.subplots(3, 1, figsize=(11, 12.5), sharex=False)
    palette = plt.cm.viridis(np.linspace(0.08, 0.92, 7))
    for axis, engine in zip(axes, engines):
        engine_data = frame[(frame["engine"] == engine) & (frame["point_count"] >= 10)]
        for color, (current, group) in zip(palette, engine_data.groupby("current_point_A")):
            group = group.sort_values("date")
            axis.plot(group["date"], group["voltage_median_V"], "o", color=color, alpha=0.38, ms=3.5)
            axis.plot(group["date"], add_filter(group, "voltage_median_V"), color=color, lw=2, label=f"{current:g} A")
        axis.set_title(f"{engine}：各典型电流点电压趋势")
        axis.set_ylabel("系统总电压 / V")
        axis.legend(ncol=6, fontsize=8, frameon=False)
    axes[-1].set_xlabel("日期")
    fig.suptitle("散点为未滤波日中位数，实线为3个代表日滚动中位数", y=1.01, fontsize=11)
    fig.tight_layout()
    fig.savefig(output, bbox_inches="tight")
    plt.close(fig)


def plot_common(frame: pd.DataFrame, output: Path) -> None:
    fig, axes = plt.subplots(2, 1, figsize=(10.5, 8.3))
    colors = {"\u4e09\u53f7\u53d1\u52a8\u673a": "#2166AC", "\u56db\u53f7\u53d1\u52a8\u673a": "#B2182B", "\u4e94\u53f7\u53d1\u52a8\u673a": "#1B7837"}
    for axis, current in zip(axes, [90, 120]):
        for engine, group in frame[(frame["current_point_A"] == current) & (frame["point_count"] >= 10)].groupby("engine", sort=False):
            group = group.sort_values("date")
            axis.plot(group["date"], group["voltage_median_V"], "o", color=colors[engine], alpha=0.35, ms=3.8)
            axis.plot(group["date"], add_filter(group, "voltage_median_V"), color=colors[engine], lw=2.3, label=engine)
        axis.set(title=f"共同工作点 {current} A", ylabel="系统总电压 / V")
        axis.legend(frameon=False, ncol=3, fontsize=8)
    axes[-1].set_xlabel("日期")
    fig.tight_layout()
    fig.savefig(output, bbox_inches="tight")
    plt.close(fig)


def write_report(inventory_frame: pd.DataFrame, summary: pd.DataFrame, output: Path, sampled_days: int) -> None:
    valid = summary[(summary["status"] == "ok") & summary["voltage_median_V"].notna()]
    coverage = valid.groupby(["engine", "current_point_A"]).agg(days=("date", "nunique"), samples=("point_count", "sum"), voltage_median_V=("voltage_median_V", "median"), voltage_std_V=("voltage_median_V", "std")).reset_index()
    lines = [
        "# \u674e\u4fca\u8c6a\u4e09\u53f0\u53d1\u52a8\u673a\u6570\u636e\u4e0e\u7535\u538b\u8d8b\u52bf\u5ba1\u8ba1",
        "",
        "## \u76f4\u63a5\u7ed3\u8bba",
        "",
        "\u8be5\u76ee\u5f55\u9002\u5408\u505a\u957f\u671f\u5b9e\u8f66\u5bf9\u6bd4，\u4f46\u4e0d\u662f“\u4e09\u4e2a\u5355\u5806”\uff1a\u5b83\u5305\u542b\u4e09\u53f0\u53cc\u5806\u71c3\u6599\u7535\u6c60\u53d1\u52a8\u673a（\u4e09\u53f7/\u56db\u53f7/\u4e94\u53f7），\u6bcf\u53f0\u90fd\u6709 A \u5806\u4e0e B \u5806\u6570\u636e\u3002\u672c\u6b21\u4e3b\u56fe\u6bd4\u8f83\u4e09\u53f0\u53d1\u52a8\u673a\u7684\u7cfb\u7edf\u603b\u7535\u538b，\u540c\u65f6\u4fdd\u7559 A/B \u5806\u5e73\u5747\u7535\u538b\u4f9b\u540e\u7eed\u5355\u5806\u4e0d\u4e00\u81f4\u6027\u5ba1\u8ba1\u3002",
        "",
        f"\u7531\u4e8e\u539f\u59cb XLSX \u5171\u7ea6 17.7 GB，\u672c\u6b21\u662f\u6570\u636e\u9009\u62e9\u9884\u5ba1\uff0c\u6bcf\u53f0\u53d1\u52a8\u673a\u6309\u5b8c\u6574\u65f6\u95f4\u8f74\u5747\u5300\u9009\u53d6 {sampled_days} \u4e2a\u4ee3\u8868\u65e5\uff0c\u6ca1\u6709\u5bf9\u5168\u90e8\u79d2\u7ea7\u6837\u672c\u505a\u6700\u7ec8\u9000\u5316\u62df\u5408\u3002",
        "",
        "## \u6570\u636e\u76ee\u5f55",
        "",
        inventory_frame.to_markdown(index=False),
        "",
        "## process_data \u811a\u672c\u7684\u4f5c\u7528",
        "",
        "- `main_data.py`：\u8bfb\u53d6\u5df2\u751f\u6210\u7684\u4e94\u53f7\u53d1\u52a8\u673a current/voltage/airflow/power pickle，\u6bcf 5 min \u53d6\u6837，\u5bfc\u51fa `Data_B.xlsx`\u3002",
        "- `main_Poly.py`：\u5bf9\u4e94\u53f7\u7684\u7535\u6d41\u2014\u529f\u7387\u5173\u7cfb\u505a\u4e09\u6b21\u591a\u9879\u5f0f\u62df\u5408，\u518d\u5bf9\u5f52\u4e00\u5316\u529f\u7387\u504f\u5dee\u4e24\u6b21\u79fb\u52a8\u5e73\u5747\u3002",
        "- `main_Rati.py`：\u5bf9\u4e09\u53f7\u7535\u6d41\u2014\u529f\u7387\u505a\u4e09\u6b21/\u4e00\u6b21\u6709\u7406\u51fd\u6570\u62df\u5408，\u5bfc\u51fa `BOL_A.xlsx`，\u5e76\u5e73\u6ed1\u529f\u7387\u504f\u5dee\u3002",
        "- `main_Rati_2.py`：\u5bf9\u4e94\u53f7\u505a\u4e8c\u6b21/\u4e8c\u6b21\u6709\u7406\u51fd\u6570\u62df\u5408\u4e0e\u504f\u5dee\u5e73\u6ed1\u3002",
        "- `main_Sine.py`：\u5bf9\u4e09\u53f7\u505a\u6b63\u5f26\u51fd\u6570\u7535\u6d41\u2014\u529f\u7387\u62df\u5408，\u7528\u4e8e\u6bd4\u8f83\u4e0d\u540c\u62df\u5408\u5f62\u5f0f\u3002",
        "- `main_SVR.py`：\u5bf9\u4e09\u53f7\u505a RBF-SVR \u7535\u6d41\u2014\u529f\u7387\u62df\u5408，\u540c\u6837\u5e73\u6ed1\u5f52\u4e00\u5316\u504f\u5dee\u3002",
        "- `main_soc.py`：\u622a\u53d6\u4e94\u53f7\u7b2c 5 \u5929\u529f\u7387/SOC，\u6bcf 5 min \u53d6\u6837\u5e76\u5bfc\u51fa `SOC_B.xlsx`\u3002",
        "- `moving_average.py`：\u63d0\u4f9b\u4e2d\u503c\u6ee4\u6ce2\u548c\u4e00\u7ef4/\u4e8c\u7ef4\u79fb\u52a8\u5e73\u5747\u5de5\u5177\u3002",
        "",
        "\u5173\u952e\u9650\u5236：\u4e0a\u8ff0\u811a\u672c\u7684\u8f93\u5165\u662f\u5df2\u7ecf\u751f\u6210\u7684 `.data` pickle，`process_data` \u76ee\u5f55\u6ca1\u6709\u627e\u5230\u4ece\u6bcf\u65e5 XLSX \u751f\u6210\u8fd9\u4e9b pickle \u7684\u5b8c\u6574\u811a\u672c；\u800c\u4e14\u73b0\u6709\u811a\u672c\u4e3b\u8981\u7528\u4e09\u53f7\u548c\u4e94\u53f7，\u56db\u53f7\u6ca1\u6709\u5bf9\u5e94 `*_fourth.data`\u3002\u56e0\u6b64\u4e0d\u80fd\u628a\u5b83\u5f53\u6210\u4e09\u53f0\u53d1\u52a8\u673a\u539f\u59cb\u6570\u636e\u7684\u5b8c\u6574\u53ef\u590d\u73b0\u5904\u7406\u94fe\u3002",
        "",
        "## \u5178\u578b\u7535\u6d41\u70b9\u53ca\u8986\u76d6",
        "",
        "\u5178\u578b\u70b9\u4e0d\u662f\u4eba\u4e3a\u6309 kW \u7b49\u95f4\u9694\u5212\u5206，\u800c\u662f\u7531\u5b9e\u8f66\u8fd0\u884c\u7535\u6d41\u76f4\u65b9\u56fe\u4e2d\u7684\u9ad8\u9891\u5e73\u53f0\u786e\u5b9a\u3002\u672c\u6b21\u6309 \u00b11 A \u9009\u53d6\u8fd0\u884c、\u65e0\u6545\u969c\u6837\u672c，\u6bcf\u65e5\u53d6\u7535\u538b\u4e2d\u4f4d\u6570\u3002",
        "",
        coverage.to_markdown(index=False, floatfmt=".3f"),
        "",
        "\u4e09\u53f0\u5171\u540c\u4e14\u9002\u5408\u6a2a\u5411\u5bf9\u6bd4\u7684\u5de5\u4f5c\u70b9\u662f **90 A \u548c 120 A**\u3002\u5176\u4f59\u5e73\u53f0\u7528\u4e8e\u5404\u53d1\u52a8\u673a\u81ea\u8eab\u7684\u7eb5\u5411\u8d8b\u52bf，\u4e0d\u5b9c\u5f3a\u884c\u505a\u4e09\u673a\u6bd4\u8f83\u3002",
        "",
        "## \u5efa\u8bae\u9009\u54ea\u4e9b\u6570\u636e",
        "",
        "1. \u82e5\u76ee\u6807\u662f\u4e09\u673a\u7edf\u4e00\u5bf9\u6bd4，\u5148\u4f7f\u7528 90 A \u548c 120 A，\u56e0\u4e3a\u8fd9\u4e24\u4e2a\u70b9\u5728\u4e09\u53f0\u53d1\u52a8\u673a\u4e2d\u90fd\u5b58\u5728\u3002",
        "2. \u82e5\u76ee\u6807\u662f\u5355\u673a\u957f\u671f\u9000\u5316，\u4f18\u5148\u9009\u62e9\u65e5\u671f\u8986\u76d6\u5b8c\u6574、\u540c\u4e00\u7535\u6d41\u70b9\u6837\u672c\u591a、\u4e14 A/B \u5806\u5b57\u6bb5\u5b8c\u6574\u7684\u65e5\u6587\u4ef6；\u4e0d\u8981\u76f4\u63a5\u62fc\u63a5\u4e09\u53f0\u53d1\u52a8\u673a\u4e3a\u4e00\u6761\u9000\u5316\u8f68\u8ff9\u3002",
        "3. \u62df\u5408\u524d\u8fd8\u9700\u8981\u6309\u6c34\u6e29、\u7a7a\u6c14\u538b\u529b/\u6d41\u91cf、Purge、\u6545\u969c、\u542f\u505c\u540e\u70ed\u7a33\u5b9a\u65f6\u95f4\u8fdb\u4e00\u6b65\u7b5b\u9009；\u672c\u6b21\u56fe\u53ea\u7528\u4e8e\u9009\u6570\u636e，\u4e0d\u662f\u6700\u7ec8\u9000\u5316\u7387\u3002",
        "4. \u56db\u53f7\u76ee\u5f55\u4e2d\u5b58\u5728\u5f88\u5c0f\u6216\u6709\u6548\u8fd0\u884c\u884c\u4e3a 0 \u7684\u65e5\u6587\u4ef6，\u5fc5\u987b\u5148\u505a\u9010\u65e5\u8d28\u91cf\u7b49\u7ea7，\u4e0d\u80fd\u53ea\u6309\u6587\u4ef6\u5b58\u5728\u5c31\u7eb3\u5165\u3002",
        "",
        "## \u56fe\u7684\u89e3\u91ca",
        "",
        "- \u201c\u539f\u59cb\u201d\u6307\u4ee3\u8868\u65e5\u5185\u6240\u6709\u7b26\u5408\u6761\u4ef6\u6837\u672c\u7684\u672a\u5e73\u6ed1\u65e5\u4e2d\u4f4d\u6570，\u4e0d\u662f\u628a\u6574\u5e74\u6570\u5343\u4e07\u4e2a\u79d2\u7ea7\u70b9\u5806\u5728\u4e00\u5f20\u56fe\u4e0a\u3002",
        "- \u201c\u6ee4\u6ce2\u201d\u662f 3 \u4e2a\u4ee3\u8868\u65e5\u7684\u5c45\u4e2d\u6eda\u52a8\u4e2d\u4f4d\u6570，\u53ea\u7528\u4e8e\u663e\u793a\u7f13\u6162\u8d8b\u52bf，\u4e0d\u4f5c\u4e3a\u7269\u7406\u8001\u5316\u7cfb\u6570\u3002",
    ]
    output.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--data-root", type=Path, default=DEFAULT_DATA_ROOT)
    parser.add_argument("--output", type=Path, default=Path("experiments/li_junhao_voltage_audit"))
    parser.add_argument("--sample-days", type=int, default=9)
    parser.add_argument("--tolerance", type=float, default=1.0)
    args = parser.parse_args()

    args.output.mkdir(parents=True, exist_ok=True)
    plot_dir = args.output / "plots"
    plot_dir.mkdir(exist_ok=True)
    records, inventory_frame = inventory(args.data_root)
    inventory_frame.to_csv(args.output / "data_inventory.csv", index=False, encoding="utf-8-sig")

    selected: list[FileRecord] = []
    for engine in ENGINE_DIRS:
        engine_records = sorted((item for item in records if item.engine == engine), key=lambda item: item.date)
        selected.extend(uniform_sample(engine_records, args.sample_days))

    summary_rows: list[dict] = []
    for index, record in enumerate(selected, start=1):
        print(f"[{index}/{len(selected)}] {record.path.name}", flush=True)
        try:
            summary_rows.extend(read_daily_summary(record, args.tolerance))
        except Exception as error:
            summary_rows.append({"engine": record.engine, "date": record.date, "file": record.path.name, "status": f"error:{type(error).__name__}"})
    summary = pd.DataFrame(summary_rows)
    summary.to_csv(args.output / "representative_day_voltage_summary.csv", index=False, encoding="utf-8-sig")
    configure_plot_style()
    plot_overall(summary, plot_dir / "three_engine_overall_voltage_trend.png")
    plot_typical(summary, plot_dir / "typical_current_voltage_trends.png")
    plot_common(summary, plot_dir / "common_90A_120A_voltage_comparison.png")
    write_report(inventory_frame, summary, args.output / "li_junhao_data_selection_report.md", args.sample_days)
    print(f"Outputs: {args.output.resolve()}")


if __name__ == "__main__":
    main()
