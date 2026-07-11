"""Inspect one raw workbook per Li Junhao engine without changing source data."""

from __future__ import annotations

from pathlib import Path
import json

import pandas as pd
from openpyxl import load_workbook


DATA_ROOT = Path(__file__).resolve().parents[3] / "2026\u674e\u4fca\u8c6a" / "\u5b66\u4f4d\u8bba\u6587" / "\u5b9e\u9a8c\u6570\u636e"
ENGINE_DIRS = {
    "\u4e09\u53f7\u53d1\u52a8\u673a": "all_data",
    "\u56db\u53f7\u53d1\u52a8\u673a": "YF20YF21-20210707-20220110",
    "\u4e94\u53f7\u53d1\u52a8\u673a": "YF24-20210908-20220208",
}


def first_valid_file(engine: str, relative: str) -> Path:
    files = [
        path
        for path in sorted((DATA_ROOT / relative).glob(f"{engine}_*.xlsx"))
        if path.stat().st_size > 1_000_000
    ]
    if not files:
        raise FileNotFoundError(engine)
    return files[0]


def inspect_workbook(engine: str, path: Path) -> tuple[list[dict], dict]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    field_rows: list[dict] = []
    overview: dict = {"engine": engine, "file": path.name, "sheets": {}}
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        iterator = worksheet.iter_rows(values_only=True)
        try:
            headers = list(next(iterator))
        except StopIteration:
            headers = []
        clean_headers = [str(value).strip() if value is not None else "" for value in headers]
        overview["sheets"][sheet_name] = {
            "rows_including_header": worksheet.max_row,
            "columns": worksheet.max_column,
            "headers": clean_headers,
        }
        for index, header in enumerate(clean_headers, start=1):
            field_rows.append(
                {
                    "engine": engine,
                    "file": path.name,
                    "sheet": sheet_name,
                    "column_index": index,
                    "field": header,
                }
            )
    workbook.close()
    return field_rows, overview


def main() -> None:
    output = Path("experiments/li_junhao_dual_stack_audit")
    output.mkdir(parents=True, exist_ok=True)
    fields: list[dict] = []
    overviews: list[dict] = []
    for engine, relative in ENGINE_DIRS.items():
        path = first_valid_file(engine, relative)
        print(f"Inspecting {path.name}", flush=True)
        engine_fields, overview = inspect_workbook(engine, path)
        fields.extend(engine_fields)
        overviews.append(overview)
    pd.DataFrame(fields).to_csv(output / "one_day_all_sheet_field_dictionary.csv", index=False, encoding="utf-8-sig")
    (output / "one_day_workbook_schema.json").write_text(json.dumps(overviews, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
